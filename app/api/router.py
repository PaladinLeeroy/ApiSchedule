from fastapi import APIRouter, HTTPException, Depends, Request, status
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime, date, timedelta
from typing import Optional, List
from fastapi.security import OAuth2PasswordRequestForm
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
import re

from app.api.database import SessionLocal
from app.api.schemas import GroupCreate, TeacherCreate, TeacherUpdate, LessonsCreate, LessonsUpdate, ScheduleCreate, RoomCreate, RoomUpdate, ScheduleTemplateCreate, ScheduleTemplateResponse, ScheduleResponse, ScheduleWithDetails, UserCreate, UserResponse, Token, TodoCreate, TodoUpdate, TodoResponse
from app.api.models import Groups, Teachers, Lessons, Schedule, Rooms, ScheduleTemplate, User, Todo
from app.api.utils import extract_year
from app.api.auth import (
    verify_password, get_password_hash, create_access_token,
    get_current_user, ACCESS_TOKEN_EXPIRE_MINUTES
)

api_router = APIRouter(prefix="/api")
html_router = APIRouter()
templates = Jinja2Templates(directory="app/templates")

# Dependency для получения сессии базы данных
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@html_router.get("/", response_class=HTMLResponse, summary="Главная страница")
async def read_items(request: Request, db: Session = Depends(get_db)):
    return templates.TemplateResponse("index.html", {"request": request, "page_title": "Главная"})

@html_router.get("/new_schedule/", response_class=HTMLResponse, summary="Создание нового расписания")
async def new_schedule(request: Request, db: Session = Depends(get_db)):
    groups = db.query(Groups).all()
    lessons = db.query(Lessons).all()
    teachers = db.query(Teachers).all()
    return templates.TemplateResponse("new_schedule.html", {
        "request": request, 
        "groups": groups, 
        "lessons": lessons, 
        "teachers": teachers,
        "page_title": "Создание расписания"
    })

@html_router.get("/ready_schedule/", response_class=HTMLResponse, summary="Страница готовых расписаний")
async def ready_schedule(request: Request, db: Session = Depends(get_db)):
    return templates.TemplateResponse("ready_schedule.html", {
        "request": request,
        "page_title": "Готовые расписания"
    })

@html_router.get("/group/", response_class=HTMLResponse, summary="Страница групп студентов")
async def group(request: Request, db: Session = Depends(get_db)):
    groups = db.query(Groups).all()
    sorted_groups = sorted(groups, key=lambda group: extract_year(group.name))
    return templates.TemplateResponse("group.html", {
        "request": request, 
        "groups": sorted_groups,
        "page_title": "Группы студентов"
    })

@html_router.get("/rooms/", response_class=HTMLResponse, summary="Страница кабинетов")
async def rooms(request: Request, db: Session = Depends(get_db)):
    return templates.TemplateResponse("rooms.html", {
        "request": request,
        "page_title": "Кабинеты"
    })

@html_router.get("/teachers/", response_class=HTMLResponse, summary="Страница преподавателей")
async def teachers(request: Request, db: Session = Depends(get_db)):
    teachers = db.query(Teachers).order_by(Teachers.name).all()
    return templates.TemplateResponse("teachers.html", {
        "request": request, 
        "teachers": teachers,
        "page_title": "Преподаватели"
    })

@html_router.get("/lessons/", response_class=HTMLResponse, summary="Страница предметов")
async def lessons(request: Request, db: Session = Depends(get_db)):
    teachers = db.query(Teachers).all()
    lessons = db.query(Lessons).join(Teachers).order_by(Lessons.name).all()
    return templates.TemplateResponse("lessons.html", {
        "request": request, 
        "teachers": teachers, 
        "lessons": lessons,
        "page_title": "Предметы"
    })

@html_router.get("/unauthorized", response_class=HTMLResponse)
async def unauthorized(request: Request):
    return templates.TemplateResponse("unauthorized.html", {
        "request": request,
        "page_title": "Авторизация"
    })

@html_router.get("/view_schedule/{template_id}", response_class=HTMLResponse)
async def view_schedule(template_id: int, request: Request, db: Session = Depends(get_db)):
    template = db.query(ScheduleTemplate).filter(ScheduleTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Расписание не найдено")
    return templates.TemplateResponse("view_schedule.html", {
        "request": request,
        "template_id": template_id,
        "template": template,
        "page_title": "Просмотр расписания"
    })

@html_router.get("/todo/", response_class=HTMLResponse, summary="ToDo List")
async def todo_page(request: Request):
    return templates.TemplateResponse("todo.html", {"request": request, "page_title": "ToDo List"})

@api_router.post("/new_group", summary="Создание группы")
async def new_group(group: GroupCreate, db: Session = Depends(get_db)):
    existing_group = db.query(Groups).filter(Groups.name == group.name).first()
    if existing_group:
        raise HTTPException(status_code=404, detail="Группа с таким именем уже существует.")
    db_group = Groups(name=group.name, type=group.type, description=group.description)
    db.add(db_group)
    db.commit()
    db.refresh(db_group)
    db.close()
    return {"message": "Group added successfully!"}

@api_router.delete("/group/{group_id}", summary="Удаление групп")
async def delete_group(group_id: int, db: Session = Depends(get_db)):
    deletes_group = db.get(Groups, group_id)
    if not deletes_group:
        raise HTTPException(status_code=404, detail="Группа не найдена")
    db.delete(deletes_group)
    db.commit()
    db.close()
    return {"ok": True}

@api_router.post("/new_teacher", summary="Создание преподавателя")
async def new_teacher(teacher: TeacherCreate, db: Session = Depends(get_db)):
    existing_teacher = db.query(Teachers).filter(Teachers.name == teacher.name).first()
    if existing_teacher:
        raise HTTPException(status_code=404, detail="Такой преподаватель уже существует.")
    db_teacher = Teachers(name=teacher.name, description=teacher.description)
    db.add(db_teacher)
    db.commit()
    db.refresh(db_teacher)
    db.close()
    return {"message": "Teacher added successfully!"}

@api_router.put("/teachers/{teacher_id}", summary="Изменение данных преподавателя")
async def update_teacher(teacher_id: int, teacher: TeacherUpdate, db: Session = Depends(get_db)):
    existing_teacher = db.get(Teachers, teacher_id)
    if not existing_teacher:
        raise HTTPException(status_code=404, detail="Преподаватель не найден")
    existing_teacher.name = teacher.name
    existing_teacher.description = teacher.description
    db.commit()
    db.refresh(existing_teacher)
    db.close()
    return {"message": "Teacher updated successfully!"}

@api_router.delete("/teacher/{teacher_id}", summary="Удаление преподавателя")
async def delete_teacher(teacher_id: int, db: Session = Depends(get_db)):
    deletes_teacher = db.get(Teachers, teacher_id)
    if not deletes_teacher:
        raise HTTPException(status_code=404, detail="Преподаватель не найден")
    db.delete(deletes_teacher)
    db.commit()
    db.close()
    return {"ok": True}

@api_router.post("/new_lesson", summary="Добавление предмета")
async def new_teacher(lesson: LessonsCreate, db: Session = Depends(get_db)):
    existing_lesson = db.query(Lessons).filter(
        Lessons.name == lesson.name,
        Lessons.teacher_id == lesson.teacher
    ).first()
    if existing_lesson:
        raise HTTPException(status_code=400, detail="Lesson with this teacher already exists.")
    db_lesson = Lessons(name=lesson.name, teacher_id=lesson.teacher)
    db.add(db_lesson)
    db.commit()
    db.refresh(db_lesson)
    db.close()
    return {"message": "Lesson added successfully!"}

@api_router.delete("/lessons/{lesson_id}", summary="Удаление предмета")
async def delete_teacher(lesson_id: int, db: Session = Depends(get_db)):
    deletes_lesson = db.get(Lessons, lesson_id)
    if not deletes_lesson:
        raise HTTPException(status_code=404, detail="Предмет не найден")
    db.delete(deletes_lesson)
    db.commit()
    db.close()
    return {"ok": True}

@api_router.put("/lessons/{lesson_id}", summary="Изменение данных предмета")
async def update_lesson(lesson_id: int, lesson: LessonsUpdate, db: Session = Depends(get_db)):
    existing_lesson = db.get(Lessons, lesson_id)
    if not existing_lesson:
        raise HTTPException(status_code=404, detail="Предмет не найден")
    
    existing_lesson_with_same_name = db.query(Lessons).filter(
        Lessons.name == lesson.name,
        Lessons.teacher_id == lesson.teacher,
        Lessons.id != lesson_id
    ).first()
    
    if existing_lesson_with_same_name:
        raise HTTPException(status_code=400, detail="Предмет с таким именем у этого преподавателя уже существует")
    
    existing_lesson.name = lesson.name
    existing_lesson.teacher_id = lesson.teacher
    db.commit()
    db.refresh(existing_lesson)
    db.close()
    return {"message": "Lesson updated successfully!"}

@api_router.get("/lessons-with-teachers", summary="Получение списка предметов с преподавателями")
async def get_lessons_with_teachers(db: Session = Depends(get_db)):
    lessons_with_teachers = db.query(Lessons, Teachers).join(Teachers).all()
    result = []
    for lesson, teacher in lessons_with_teachers:
        result.append({
            "id": lesson.id,
            "name": lesson.name,
            "teacher_id": teacher.id,
            "teacher_name": teacher.name
        })
    return result

@api_router.post("/schedule-template", response_model=ScheduleTemplateResponse)
def create_schedule_template(
    template: ScheduleTemplateCreate,
    db: Session = Depends(get_db)
):
    # Проверяем, существует ли уже расписание для данного типа группы в указанный период
    existing_template = db.query(ScheduleTemplate).filter(
        ScheduleTemplate.group_type == template.group_type,
        ScheduleTemplate.date_start <= template.date_end,
        ScheduleTemplate.date_end >= template.date_start
    ).first()
    
    if existing_template:
        raise HTTPException(
            status_code=400,
            detail="Для данного типа группы уже существует расписание в указанный период"
        )
    
    db_template = ScheduleTemplate(
        name=template.name,
        group_type=template.group_type,
        date_start=template.date_start,
        date_end=template.date_end,
        schedule_type=template.schedule_type
    )
    db.add(db_template)
    db.commit()
    db.refresh(db_template)
    return db_template

@api_router.post("/schedule", response_model=ScheduleResponse)
async def create_schedule(schedule: ScheduleCreate, db: Session = Depends(get_db)):
    # Проверяем существование шаблона
    template = db.query(ScheduleTemplate).filter(ScheduleTemplate.id == schedule.template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Шаблон расписания не найден")
    
    # Проверяем, что дата находится в пределах шаблона
    if not (template.date_start <= schedule.date <= template.date_end):
        raise HTTPException(status_code=400, detail="Дата выходит за пределы шаблона")
    
    # Проверяем существование группы
    group = db.query(Groups).filter(Groups.id == schedule.group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Группа не найдена")
    
    # Проверяем, что тип группы соответствует шаблону
    if group.type != template.group_type:
        raise HTTPException(status_code=400, detail="Тип группы не соответствует шаблону")
    
    # Проверяем существование занятия
    lesson = db.query(Lessons).filter(Lessons.id == schedule.lesson_id).first()
    if not lesson:
        raise HTTPException(status_code=404, detail="Занятие не найдено")
    
    # Проверяем существование преподавателя
    teacher = db.query(Teachers).filter(Teachers.id == schedule.teacher_id).first()
    if not teacher:
        raise HTTPException(status_code=404, detail="Преподаватель не найден")
    
    # Проверяем существование кабинета
    room = db.query(Rooms).filter(Rooms.id == schedule.room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="Кабинет не найден")
    
    # Создаем новое расписание
    db_schedule = Schedule(
        template_id=schedule.template_id,
        date=schedule.date,
        lesson_number=schedule.lesson_number,
        group_id=schedule.group_id,
        lesson_id=schedule.lesson_id,
        teacher_id=schedule.teacher_id,
        room_id=schedule.room_id,
        is_above_line=schedule.is_above_line,
        lesson_type=schedule.lesson_type,
        day_of_week=schedule.day_of_week
    )
    
    db.add(db_schedule)
    db.commit()
    db.refresh(db_schedule)
    
    return db_schedule

@api_router.get("/groups", summary="Получение списка групп")
async def get_groups(db: Session = Depends(get_db)):
    groups = db.query(Groups).all()
    return [{"id": group.id, "name": group.name, "type": group.type, "description": group.description} for group in groups]

@api_router.post("/groups", summary="Создание новой группы через API")
async def create_group_api(group: GroupCreate, db: Session = Depends(get_db)):
    existing_group = db.query(Groups).filter(Groups.name == group.name).first()
    if existing_group:
        raise HTTPException(status_code=400, detail="Группа с таким именем уже существует.")
    db_group = Groups(name=group.name, type=group.type, description=group.description)
    db.add(db_group)
    db.commit()
    db.refresh(db_group)
    return {"message": "Group added successfully!", "id": db_group.id}

@api_router.get("/rooms", summary="Получение списка кабинетов")
async def get_rooms(db: Session = Depends(get_db)):
    rooms = db.query(Rooms).all()
    return [{"id": room.id, "number": room.number, "capacity": room.capacity, "description": room.description} for room in rooms]

@api_router.post("/rooms", summary="Создание нового кабинета")
async def create_room(room: RoomCreate, db: Session = Depends(get_db)):
    existing_room = db.query(Rooms).filter(Rooms.number == room.number).first()
    if existing_room:
        raise HTTPException(status_code=400, detail="Кабинет с таким номером уже существует")
    
    db_room = Rooms(
        number=room.number,
        capacity=room.capacity,
        description=room.description
    )
    db.add(db_room)
    db.commit()
    db.refresh(db_room)
    return {"message": "Room added successfully!"}

@api_router.put("/rooms/{room_id}", summary="Обновление информации о кабинете")
async def update_room(room_id: int, room: RoomUpdate, db: Session = Depends(get_db)):
    db_room = db.query(Rooms).filter(Rooms.id == room_id).first()
    if not db_room:
        raise HTTPException(status_code=404, detail="Кабинет не найден")
    
    if room.number:
        existing_room = db.query(Rooms).filter(Rooms.number == room.number, Rooms.id != room_id).first()
        if existing_room:
            raise HTTPException(status_code=400, detail="Кабинет с таким номером уже существует")
        db_room.number = room.number
    
    if room.capacity is not None:
        db_room.capacity = room.capacity
    
    if room.description is not None:
        db_room.description = room.description
    
    db.commit()
    db.refresh(db_room)
    return {"message": "Room updated successfully!"}

@api_router.delete("/rooms/{room_id}", summary="Удаление кабинета")
async def delete_room(room_id: int, db: Session = Depends(get_db)):
    db_room = db.query(Rooms).filter(Rooms.id == room_id).first()
    if not db_room:
        raise HTTPException(status_code=404, detail="Кабинет не найден")
    
    db.delete(db_room)
    db.commit()
    return {"message": "Room deleted successfully!"}

@api_router.delete("/groups/{group_id}", summary="Удаление группы")
async def delete_group(group_id: int, db: Session = Depends(get_db)):
    db_group = db.query(Groups).filter(Groups.id == group_id).first()
    if not db_group:
        raise HTTPException(status_code=404, detail="Group not found")
    db.delete(db_group)
    db.commit()
    return {"message": "Group deleted successfully!"}

@api_router.put("/groups/{group_id}", summary="Обновление группы")
async def update_group(group_id: int, group: GroupCreate, db: Session = Depends(get_db)):
    db_group = db.query(Groups).filter(Groups.id == group_id).first()
    if not db_group:
        raise HTTPException(status_code=404, detail="Group not found")
    # Проверяем, не существует ли уже группа с таким именем
    existing_group = db.query(Groups).filter(Groups.name == group.name, Groups.id != group_id).first()
    if existing_group:
        raise HTTPException(status_code=400, detail="Group with this name already exists")
    db_group.name = group.name
    db_group.type = group.type
    db_group.description = group.description
    db.commit()
    db.refresh(db_group)
    return {"message": "Group updated successfully!"}

@api_router.get("/schedule")
async def get_schedule(
    date_start: str,
    date_end: str,
    group_type: str,
    db: Session = Depends(get_db)
):
    try:
        # Получаем все группы указанного типа
        groups = db.query(Groups).filter(Groups.type == group_type).all()
        if not groups:
            return []
        
        group_ids = [group.id for group in groups]
        
        # Получаем расписание для выбранных групп и дат
        schedule = db.query(Schedule).filter(
            Schedule.date >= date_start,
            Schedule.date <= date_end,
            Schedule.group_id.in_(group_ids)
        ).all()
        
        return schedule
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/schedule-templates", response_model=List[ScheduleTemplateResponse])
def get_schedule_templates(
    group_type: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(ScheduleTemplate)
    if group_type:
        query = query.filter(ScheduleTemplate.group_type == group_type)
    return query.all()

@api_router.get("/schedule-templates/{template_id}", response_model=ScheduleTemplateResponse)
def get_schedule_template(template_id: int, db: Session = Depends(get_db)):
    template = db.query(ScheduleTemplate).filter(ScheduleTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Шаблон расписания не найден")
    return template

@api_router.get("/schedules", response_model=List[ScheduleWithDetails])
def get_schedules(
    template_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    query = db.query(Schedule).join(
        Schedule.lesson
    ).join(
        Schedule.teacher
    ).join(
        Schedule.room
    ).join(
        Schedule.group
    ).join(
        Schedule.template
    )
    if template_id:
        query = query.filter(Schedule.template_id == template_id)
    return query.all()

@api_router.delete("/schedule-templates/{template_id}", summary="Удаление шаблона расписания")
async def delete_schedule_template(template_id: int, db: Session = Depends(get_db)):
    template = db.query(ScheduleTemplate).filter(ScheduleTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Шаблон расписания не найден")
    
    # Удаляем все связанные расписания
    db.query(Schedule).filter(Schedule.template_id == template_id).delete()
    
    # Удаляем сам шаблон
    db.delete(template)
    db.commit()
    return {"message": "Шаблон расписания успешно удален"}

@api_router.post("/token", response_model=Token)
async def login_for_access_token(
    form_data: OAuth2PasswordRequestForm = Depends(),
    db: Session = Depends(get_db)
):
    user = db.query(User).filter(User.username == form_data.username).first()
    if not user or not verify_password(form_data.password, user.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Неверное имя пользователя или пароль",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username, "role": user.role},
        expires_delta=access_token_expires
    )
    
    return JSONResponse(
        content={
            "access_token": access_token,
            "token_type": "bearer",
            "role": user.role
        },
        status_code=status.HTTP_200_OK
    )

@api_router.post("/register", response_model=UserResponse)
async def register_user(user: UserCreate, db: Session = Depends(get_db)):
    # Проверяем, существует ли пользователь
    db_user = db.query(User).filter(User.username == user.username).first()
    if db_user:
        raise HTTPException(
            status_code=400,
            detail="Username already registered"
        )
    
    # Создаем нового пользователя
    hashed_password = get_password_hash(user.password)
    db_user = User(
        username=user.username,
        password_hash=hashed_password,
        role=user.role
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user

@api_router.get("/users/me", response_model=UserResponse)
async def read_users_me(current_user: User = Depends(get_current_user)):
    return current_user

@api_router.get("/v1/schedule/group/{group_id}", 
    summary="Получение расписания для конкретной группы",
    description="Возвращает расписание для указанной группы на указанный период")
async def get_group_schedule(
    group_id: int,
    date_start: date,
    date_end: date,
    db: Session = Depends(get_db)
):
    try:
        # Проверяем существование группы
        group = db.query(Groups).filter(Groups.id == group_id).first()
        if not group:
            raise HTTPException(status_code=404, detail="Группа не найдена")

        # Получаем расписание для группы
        schedule = db.query(Schedule).filter(
            Schedule.group_id == group_id,
            Schedule.date >= date_start,
            Schedule.date <= date_end
        ).join(
            Schedule.lesson
        ).join(
            Schedule.teacher
        ).join(
            Schedule.room
        ).all()

        # Формируем ответ
        result = []
        for item in schedule:
            result.append({
                "date": item.date.isoformat(),
                "day_of_week": item.day_of_week,
                "lesson_number": item.lesson_number,
                "lesson": {
                    "id": item.lesson.id,
                    "name": item.lesson.name
                },
                "teacher": {
                    "id": item.teacher.id,
                    "name": item.teacher.name
                },
                "room": {
                    "id": item.room.id,
                    "number": item.room.number
                },
                "is_above_line": item.is_above_line,
                "lesson_type": item.lesson_type
            })

        return {
            "group": {
                "id": group.id,
                "name": group.name,
                "type": group.type
            },
            "schedule": result
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/v1/schedule/teacher/{teacher_id}",
    summary="Получение расписания для конкретного преподавателя",
    description="Возвращает расписание для указанного преподавателя на указанный период")
async def get_teacher_schedule(
    teacher_id: int,
    date_start: date,
    date_end: date,
    db: Session = Depends(get_db)
):
    try:
        # Проверяем существование преподавателя
        teacher = db.query(Teachers).filter(Teachers.id == teacher_id).first()
        if not teacher:
            raise HTTPException(status_code=404, detail="Преподаватель не найден")

        # Получаем расписание для преподавателя
        schedule = db.query(Schedule).filter(
            Schedule.teacher_id == teacher_id,
            Schedule.date >= date_start,
            Schedule.date <= date_end
        ).join(
            Schedule.lesson
        ).join(
            Schedule.group
        ).join(
            Schedule.room
        ).all()

        # Формируем ответ
        result = []
        for item in schedule:
            result.append({
                "date": item.date.isoformat(),
                "day_of_week": item.day_of_week,
                "lesson_number": item.lesson_number,
                "lesson": {
                    "id": item.lesson.id,
                    "name": item.lesson.name
                },
                "group": {
                    "id": item.group.id,
                    "name": item.group.name,
                    "type": item.group.type
                },
                "room": {
                    "id": item.room.id,
                    "number": item.room.number
                },
                "is_above_line": item.is_above_line,
                "lesson_type": item.lesson_type
            })

        return {
            "teacher": {
                "id": teacher.id,
                "name": teacher.name
            },
            "schedule": result
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/v1/schedule/group/{group_id}/week",
    summary="Получение расписания группы на текущую неделю",
    description="Возвращает расписание для указанной группы на текущую неделю")
async def get_group_week_schedule(
    group_id: int,
    db: Session = Depends(get_db)
):
    try:
        # Проверяем существование группы
        group = db.query(Groups).filter(Groups.id == group_id).first()
        if not group:
            raise HTTPException(status_code=404, detail="Группа не найдена")

        # Получаем даты начала и конца текущей недели
        today = date.today()
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)

        # Получаем расписание для группы
        schedule = db.query(Schedule).filter(
            Schedule.group_id == group_id,
            Schedule.date >= start_of_week,
            Schedule.date <= end_of_week
        ).join(
            Schedule.lesson
        ).join(
            Schedule.teacher
        ).join(
            Schedule.room
        ).order_by(
            Schedule.date,
            Schedule.lesson_number
        ).all()

        # Формируем ответ
        result = []
        for item in schedule:
            result.append({
                "date": item.date.isoformat(),
                "day_of_week": item.day_of_week,
                "lesson_number": item.lesson_number,
                "lesson": {
                    "id": item.lesson.id,
                    "name": item.lesson.name
                },
                "teacher": {
                    "id": item.teacher.id,
                    "name": item.teacher.name
                },
                "room": {
                    "id": item.room.id,
                    "number": item.room.number
                },
                "is_above_line": item.is_above_line,
                "lesson_type": item.lesson_type
            })

        return {
            "group": {
                "id": group.id,
                "name": group.name,
                "type": group.type
            },
            "week_start": start_of_week.isoformat(),
            "week_end": end_of_week.isoformat(),
            "schedule": result
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/v1/schedule/teacher/{teacher_id}/week",
    summary="Получение расписания преподавателя на текущую неделю",
    description="Возвращает расписание для указанного преподавателя на текущую неделю")
async def get_teacher_week_schedule(
    teacher_id: int,
    db: Session = Depends(get_db)
):
    try:
        # Проверяем существование преподавателя
        teacher = db.query(Teachers).filter(Teachers.id == teacher_id).first()
        if not teacher:
            raise HTTPException(status_code=404, detail="Преподаватель не найден")

        # Получаем даты начала и конца текущей недели
        today = date.today()
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)

        # Получаем расписание для преподавателя
        schedule = db.query(Schedule).filter(
            Schedule.teacher_id == teacher_id,
            Schedule.date >= start_of_week,
            Schedule.date <= end_of_week
        ).join(
            Schedule.lesson
        ).join(
            Schedule.group
        ).join(
            Schedule.room
        ).order_by(
            Schedule.date,
            Schedule.lesson_number
        ).all()

        # Формируем ответ
        result = []
        for item in schedule:
            result.append({
                "date": item.date.isoformat(),
                "day_of_week": item.day_of_week,
                "lesson_number": item.lesson_number,
                "lesson": {
                    "id": item.lesson.id,
                    "name": item.lesson.name
                },
                "group": {
                    "id": item.group.id,
                    "name": item.group.name,
                    "type": item.group.type
                },
                "room": {
                    "id": item.room.id,
                    "number": item.room.number
                },
                "is_above_line": item.is_above_line,
                "lesson_type": item.lesson_type
            })

        return {
            "teacher": {
                "id": teacher.id,
                "name": teacher.name
            },
            "week_start": start_of_week.isoformat(),
            "week_end": end_of_week.isoformat(),
            "schedule": result
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/schedule/{schedule_id}", summary="Удаление занятия из расписания")
async def delete_schedule(schedule_id: int, db: Session = Depends(get_db)):
    schedule = db.get(Schedule, schedule_id)
    if not schedule:
        raise HTTPException(status_code=404, detail="Занятие не найдено")
    db.delete(schedule)
    db.commit()
    db.close()
    return {"ok": True}

@api_router.put("/schedule/{schedule_id}", summary="Обновление занятия в расписании")
async def update_schedule(schedule_id: int, schedule: ScheduleCreate, db: Session = Depends(get_db)):
    db_schedule = db.get(Schedule, schedule_id)
    if not db_schedule:
        raise HTTPException(status_code=404, detail="Занятие не найдено")
    for field, value in schedule.dict().items():
        setattr(db_schedule, field, value)
    db.commit()
    db.refresh(db_schedule)
    return db_schedule

@api_router.get("/schedule-templates/{template_id}/export_excel", summary="Экспорт расписания в Excel")
async def export_schedule_excel(template_id: int, db: Session = Depends(get_db)):
    template = db.query(ScheduleTemplate).filter(ScheduleTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Шаблон расписания не найден")
    groups = db.query(Groups).filter(Groups.type == template.group_type).all()
    lessons = db.query(Schedule).filter(Schedule.template_id == template_id).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Расписание"
    # Шапка
    header = ["День", "№ пары", "Над чертой/Под чертой"] + [g.name for g in groups]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="000000")
    fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_blue = PatternFill(start_color="E3F0FF", end_color="E3F0FF", fill_type="solid")
    # Дни/даты
    daysOfWeek = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']
    if template.is_full_semester:
        days = daysOfWeek
        date_list = [None]*6
    else:
        start = template.date_start
        end = template.date_end
        days = [(start + timedelta(days=i)).isoformat() for i in range((end-start).days+1)]
        date_list = [start + timedelta(days=i) for i in range((end-start).days+1)]
    row_idx = 2
    day_row_indices = []  # Для объединения ячеек дат
    for day_idx, day in enumerate(days):
        day_start_idx = row_idx
        # Формируем подпись: дата + день недели
        if template.is_full_semester:
            date_label = daysOfWeek[day_idx]
        else:
            date_label = f"{day}\n{daysOfWeek[day_idx % 6]}"
        
        for lesson_num in range(1, 8):
            # Над чертой
            row_nad = [date_label if lesson_num == 1 else "", lesson_num, "Над чертой"]
            # Под чертой
            row_pod = ["", "", "Под чертой"]
            for group in groups:
                # Над чертой
                if template.is_full_semester:
                    lessons_nad = [l for l in lessons if l.group_id==group.id and l.lesson_number==lesson_num and l.day_of_week==day_idx+1 and l.is_above_line]
                else:
                    date_str = days[day_idx]
                    lessons_nad = [l for l in lessons if l.group_id==group.id and l.lesson_number==lesson_num and str(l.date)==date_str and l.is_above_line]
                cell_nad = ''
                if lessons_nad:
                    l = lessons_nad[0]
                    cell_nad = f"{l.lesson.name}\n{l.teacher.name}\nАуд. {l.room.number}"
                row_nad.append(cell_nad)
                # Под чертой
                if template.is_full_semester:
                    lessons_pod = [l for l in lessons if l.group_id==group.id and l.lesson_number==lesson_num and l.day_of_week==day_idx+1 and not l.is_above_line]
                else:
                    date_str = days[day_idx]
                    lessons_pod = [l for l in lessons if l.group_id==group.id and l.lesson_number==lesson_num and str(l.date)==date_str and not l.is_above_line]
                cell_pod = ''
                if lessons_pod:
                    l = lessons_pod[0]
                    cell_pod = f"{l.lesson.name}\n{l.teacher.name}\nАуд. {l.room.number}"
                row_pod.append(cell_pod)
            ws.append(row_nad)
            ws.append(row_pod)
            # Объединяем ячейки для номера пары
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx+1, end_column=2)
            # Высота строк (одинаковая)
            ws.row_dimensions[row_idx].height = 45
            ws.row_dimensions[row_idx+1].height = 45
            # Цвет для строки "Под чертой"
            for col in range(1, 4+len(groups)):
                ws.cell(row=row_idx+1, column=col).fill = fill_gray
            # Цвет для "Над чертой" (чередование)
            if (lesson_num % 2) == 0:
                for col in range(1, 4+len(groups)):
                    ws.cell(row=row_idx, column=col).fill = fill_blue
            row_idx += 2
        day_end_idx = row_idx - 1
        day_row_indices.append((day_start_idx, day_end_idx))
    
    # Объединяем ячейки для дат
    for start, end in day_row_indices:
        ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
    
    # Стилизация
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = ws[2][0]
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 16
    for col in range(4, 4+len(groups)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 28
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    group_type = template.group_type if template.group_type else 'Unknown'
    date_start = template.date_start.strftime('%Y-%m-%d') if template.date_start else ''
    date_end = template.date_end.strftime('%Y-%m-%d') if template.date_end else ''
    file_title = f"Raspisanie_{group_type}_{date_start}_{date_end}.xlsx".replace(' ', '_').replace(':', '')
    def safe_filename(s):
        return re.sub(r'[^A-Za-z0-9_.-]', '_', s)
    filename = safe_filename(file_title)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})

@api_router.get("/schedule-templates/{template_id}/export_pdf", summary="Экспорт расписания в PDF")
async def export_schedule_pdf(template_id: int, db: Session = Depends(get_db)):
    # Регистрируем шрифт DejaVuSans
    font_path = os.path.join("app", "static", "webfonts", "DejaVuSans.ttf")
    pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))
    bold_font_path = os.path.join("app", "static", "webfonts", "DejaVuSans-Bold.ttf")
    pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", bold_font_path))
    template = db.query(ScheduleTemplate).filter(ScheduleTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Шаблон расписания не найден")
    groups = db.query(Groups).filter(Groups.type == template.group_type).all()
    lessons = db.query(Schedule).filter(Schedule.template_id == template_id).all()
    daysOfWeek = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']
    header = ["День", "№ пары", "Над чертой/\nПод чертой"] + [g.name for g in groups]
    data = [header]
    # Формируем список дней и дат
    if template.is_full_semester:
        days = daysOfWeek
        date_list = [None]*6
    else:
        start = template.date_start
        end = template.date_end
        days = [(start + timedelta(days=i)).isoformat() for i in range((end-start).days+1)]
        date_list = [start + timedelta(days=i) for i in range((end-start).days+1)]
    styleSheet = getSampleStyleSheet()
    # Формируем строки таблицы
    day_row_indices = []  # Для объединения ячеек дат
    def vertical_text(s):
        return '\n'.join(list(s.replace('-', '–')))
    for day_idx, day in enumerate(days):
        day_start_idx = len(data)
        # Формируем подпись: дата + день недели
        if template.is_full_semester:
            date_label = daysOfWeek[day_idx]
        else:
            date_label = f"{day}\n{daysOfWeek[day_idx % 6]}"
        # Вертикальный текст для даты и дня недели
        date_label_vertical = '\n'.join([vertical_text(part) for part in date_label.split('\n')])
        for lesson_num in range(1, 8):
            row_nad = [date_label_vertical if lesson_num == 1 else "", lesson_num, "Над чертой"]
            row_pod = ["", "", "Под чертой"]
            for group in groups:
                # Над чертой
                if template.is_full_semester:
                    lessons_nad = [l for l in lessons if l.group_id==group.id and l.lesson_number==lesson_num and l.day_of_week==day_idx+1 and l.is_above_line]
                else:
                    date_str = days[day_idx]
                    lessons_nad = [l for l in lessons if l.group_id==group.id and l.lesson_number==lesson_num and str(l.date)==date_str and l.is_above_line]
                cell_nad = ''
                if lessons_nad:
                    l = lessons_nad[0]
                    cell_nad = f"{l.lesson.name}\n{l.teacher.name}\nАуд. {l.room.number}"
                row_nad.append(cell_nad)
                # Под чертой
                if template.is_full_semester:
                    lessons_pod = [l for l in lessons if l.group_id==group.id and l.lesson_number==lesson_num and l.day_of_week==day_idx+1 and not l.is_above_line]
                else:
                    date_str = days[day_idx]
                    lessons_pod = [l for l in lessons if l.group_id==group.id and l.lesson_number==lesson_num and str(l.date)==date_str and not l.is_above_line]
                cell_pod = ''
                if lessons_pod:
                    l = lessons_pod[0]
                    cell_pod = f"{l.lesson.name}\n{l.teacher.name}\nАуд. {l.room.number}"
                row_pod.append(cell_pod)
            data.append(row_nad)
            data.append(row_pod)
        day_end_idx = len(data) - 1
        day_row_indices.append((day_start_idx, day_end_idx))
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=10, leftMargin=10, topMargin=10, bottomMargin=10)
    styleSheet = getSampleStyleSheet()
    styleSheet['Title'].fontName = 'DejaVuSans'
    elements = []
    # Формируем красивый заголовок и имя файла
    group_type = template.group_type if template.group_type else 'Unknown'
    date_start = template.date_start.strftime('%Y-%m-%d') if template.date_start else ''
    date_end = template.date_end.strftime('%Y-%m-%d') if template.date_end else ''
    title_text = f"Расписание: {group_type} с {date_start} по {date_end}"
    file_title = f"Raspisanie_{group_type}_{date_start}_{date_end}.pdf".replace(' ', '_').replace(':', '')

    elements.append(Paragraph(title_text, styleSheet['Title']))
    elements.append(Spacer(1, 12))
    # Растянуть таблицу на всю ширину страницы
    page_width = landscape(A4)[0] - 20  # минус отступы
    ncols = len(header)
    # ширина: день, № пары, группы...
    colWidths = [40, 32, 60] + [(page_width-40-32-60)//(ncols-3)]*(ncols-3)
    table = Table(data, repeatRows=1, rowHeights=[34]*len(data), colWidths=colWidths)
    style = TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'DejaVuSans'),
        ('FONTSIZE', (0,0), (-1,-1), 7),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BACKGROUND', (0,0), (-1,0), colors.lightblue),
        ('TEXTCOLOR', (0,0), (-1,0), colors.black),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
    ])
    for i, row in enumerate(data[1:], start=1):
        if (i % 2) == 0:
            style.add('BACKGROUND', (0,i), (-1,i), colors.whitesmoke)
        elif (i % 4) < 2:
            style.add('BACKGROUND', (0,i), (-1,i), colors.HexColor('#e3f0ff'))
    for start, end in day_row_indices:
        style.add('SPAN', (0, start), (0, end))  # дата
    # Объединение ячеек с номером пары
    for i in range(1, len(data), 2):
        style.add('SPAN', (1, i), (1, i+1))
    table.setStyle(style)
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    def safe_filename(s):
        return re.sub(r'[^A-Za-z0-9_.-]', '_', s)
    filename = safe_filename(file_title)
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})

# Todo API endpoints
@api_router.post("/todos", response_model=TodoResponse, summary="Создание новой задачи")
async def create_todo(todo: TodoCreate, db: Session = Depends(get_db)):
    db_todo = Todo(
        text=todo.text,
        completed=todo.completed
    )
    db.add(db_todo)
    db.commit()
    db.refresh(db_todo)
    return db_todo

@api_router.get("/todos", response_model=List[TodoResponse], summary="Получение всех задач")
async def get_todos(db: Session = Depends(get_db)):
    todos = db.query(Todo).order_by(Todo.created_at.desc()).all()
    return todos

@api_router.get("/todos/{todo_id}", response_model=TodoResponse, summary="Получение конкретной задачи")
async def get_todo(todo_id: int, db: Session = Depends(get_db)):
    todo = db.query(Todo).filter(Todo.id == todo_id).first()
    if not todo:
        raise HTTPException(status_code=404, detail="Задача не найдена")
    return todo

@api_router.put("/todos/{todo_id}", response_model=TodoResponse, summary="Обновление задачи")
async def update_todo(todo_id: int, todo_update: TodoUpdate, db: Session = Depends(get_db)):
    todo = db.query(Todo).filter(Todo.id == todo_id).first()
    if not todo:
        raise HTTPException(status_code=404, detail="Задача не найдена")
    
    if todo_update.text is not None:
        todo.text = todo_update.text
    if todo_update.completed is not None:
        todo.completed = todo_update.completed
    
    db.commit()
    db.refresh(todo)
    return todo

@api_router.delete("/todos/{todo_id}", summary="Удаление задачи")
async def delete_todo(todo_id: int, db: Session = Depends(get_db)):
    todo = db.query(Todo).filter(Todo.id == todo_id).first()
    if not todo:
        raise HTTPException(status_code=404, detail="Задача не найдена")
    
    db.delete(todo)
    db.commit()
    return {"message": "Задача успешно удалена"} 